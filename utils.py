import time
import sys
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from openpyxl.utils.exceptions import InvalidFileException
import json

def save_screenshot(sa_instr, sa_name="SpectrumAnalyzer"):

    save_to_instr = input("📸 Do you want to save the screenshot to your Instrument? (y/n): ").strip().lower()
    if save_to_instr != "y":
        return

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder_on_instr = r"D:\GA monitor"
    filename = f"{sa_name}_{now}.png"
    full_instr_path = f"{folder_on_instr}\\{filename}"

    try:
        sa_instr.write(f':MMEM:STOR:SCR "{full_instr_path}"')
        slow_print(f"Screenshot saved on instrument as {full_instr_path}.")
    except Exception as e:
        slow_print(f"❌ Failed to save screenshot on instrument: {e}")
        return

    save_to_pc = input("💾 Do you want to transfer the screenshot to your PC? (y/n): ").strip().lower()
    if save_to_pc != "y":
        return

    try:
        save_folder = r"D:\GA monitor"
        os.makedirs(save_folder, exist_ok=True)
        save_path = os.path.join(save_folder, filename)

        slow_print("Transferring screenshot to PC...")
        screenshot_data = sa_instr.query_binary_values(f':MMEM:DATA? "{full_instr_path}"', datatype='B')
        with open(save_path, "wb") as f:
            f.write(bytearray(screenshot_data))
        slow_print(f"✅ Screenshot saved to PC at: {save_path}")
    except Exception as e:
        slow_print(f"❌ Failed to transfer screenshot to PC: {e}")


def format_unit(key, value):
    """Format frequency and power values into readable units."""
    try:
        value = float(value)
    except:
        return str(value)

    if "Frequency" in key and "Hz" in key:
        if value >= 1e9:
            return f"{value / 1e9:.3f} GHz"
        elif value >= 1e6:
            return f"{value / 1e6:.3f} MHz"
        elif value >= 1e3:
            return f"{value / 1e3:.3f} kHz"
        else:
            return f"{value:.0f} Hz"
    elif "Power" in key and "dBm" in key:
        return f"{value:.2f} dBm"

    return str(value)

def export_all_results(result_data, filename_prefix="Test_Result"):
    export_folder = "results"
    os.makedirs(export_folder, exist_ok=True)

    # Ask user for Excel export
    save_excel = input("💾 Do you want to save results to Excel? (y/n): ").strip().lower()
    if save_excel == 'y':
        today_str = datetime.now().strftime('%Y%m%d')
        excel_filename = f"{filename_prefix}_{today_str}.xlsx"
        excel_path = os.path.join(export_folder, excel_filename)

        try:
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
        except InvalidFileException:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([])
        ws.append(["Timestamp", timestamp])
        ws.append([])

        if isinstance(result_data, dict):
            ws.append(["Parameter", "Value"])
            for key, value in result_data.items():
                ws.append([key, value])
        elif isinstance(result_data, list):
            if all(isinstance(item, dict) for item in result_data):
                headers = list(result_data[0].keys())
                ws.append(headers)
                for item in result_data:
                    ws.append([item.get(h, "") for h in headers])
            else:
                ws.append(["Index", "Value"])
                for idx, item in enumerate(result_data, start=1):
                    ws.append([idx, str(item)])
        else:
            ws.append(["Raw Output"])
            ws.append([str(result_data)])

        wb.save(excel_path)
        print(f"✅ Results saved to Excel: {excel_path}")

    # Ask user for JSON export
    save_json = input("💾 Do you want to save results to JSON? (y/n): ").strip().lower()
    if save_json == 'y':
        today = datetime.now().strftime("%Y%m%d")
        json_filename = f"{filename_prefix}_{today}.json"
        json_path = os.path.join(export_folder, json_filename)

        if os.path.exists(json_path):
            try:
                with open(json_path, "r") as f:
                    log = json.load(f)
            except json.JSONDecodeError:
                log = []
        else:
            log = []

        log.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "data": result_data
        })

        with open(json_path, "w") as f:
            json.dump(log, f, indent=4)

        print(f"✅ Results saved to JSON log: {json_path}")


def parse_frequency(freq_input):
    freq_input = freq_input.strip().lower().replace(" ", "")  # e.g., '1GHz' → '1ghz'

    multipliers = {
        'ghz': 1e9,
        'mhz': 1e6,
        'khz': 1e3,
        'hz': 1,
    }

    for unit in multipliers:
        if freq_input.endswith(unit):
            try:
                value = float(freq_input.replace(unit, ""))
                return value * multipliers[unit]
            except ValueError:
                raise ValueError("Invalid numeric value in frequency input.")

    # If no unit is provided, assume Hz
    try:
        return float(freq_input)
    except ValueError:
        raise ValueError("Frequency format not recognized.")

#fucntion for printing values with delay

SLOW_MODE = True # 🔁 Toggle this to False for instant printing

def slow_print(text, delay=0.05, char_delay=0.05):
    if not SLOW_MODE:
        print(text)
        return

    for char in text:
        print(char, end='', flush=True)
        time.sleep(char_delay)
    print()
    time.sleep(delay)